#!/usr/bin/env python3
"""
Download image links from an XLSX column (default: E / "Image Link") for QA checking.

Default behavior:
- Reads rows from 2 onward
- Includes only rows where key column B has a value
- Reads image URL from column E
- Saves images into downloads/image_check
- Writes a CSV-like report with success/fail status
"""

import argparse
import csv
import mimetypes
import os
import re
import time
from pathlib import Path
from urllib.parse import urlparse
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "sources", "SAN NENG.xlsx")
DEFAULT_OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "downloads", "image_check")
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) SAN-NENG-Image-Checker/1.0"


def is_empty(value):
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def parse_args():
    parser = argparse.ArgumentParser(description="Download image URLs from XLSX column")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Path to XLSX file")
    parser.add_argument("--sheet", default="", help="Sheet name (default: active sheet)")
    parser.add_argument("--column", default="E", help="Image URL column letter (default: E)")
    parser.add_argument("--key-column", default="B", help="Only include rows where this column has value (default: B)")
    parser.add_argument("--min-row", type=int, default=2, help="First row to scan (default: 2)")
    parser.add_argument("--out-dir", default=DEFAULT_OUTPUT_DIR, help="Directory to save images")
    parser.add_argument("--timeout", type=int, default=25, help="HTTP timeout in seconds (default: 25)")
    parser.add_argument("--retries", type=int, default=2, help="Retries per URL after first attempt (default: 2)")
    parser.add_argument("--sleep", type=float, default=0.0, help="Sleep seconds between downloads (default: 0)")
    parser.add_argument("--limit", type=int, default=0, help="Maximum images to attempt (default: 0 = all)")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite files if already present")
    return parser.parse_args()


def sanitize_filename(value):
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip())
    return text.strip("._-") or "image"


def infer_extension(url, content_type):
    parsed = urlparse(url)
    ext = Path(parsed.path).suffix.lower()
    if ext in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tif", ".tiff", ".svg", ".avif"}:
        return ext

    if content_type:
        clean_type = content_type.split(";")[0].strip().lower()
        guessed = mimetypes.guess_extension(clean_type)
        if guessed:
            return guessed

    return ".jpg"


def read_bytes(url, timeout):
    request = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(request, timeout=timeout) as response:
        data = response.read()
        content_type = response.headers.get("Content-Type", "")
        return data, content_type


def resolve_cell_url(cell):
    value = cell.value
    if isinstance(value, str) and value.strip().lower().startswith(("http://", "https://")):
        return value.strip()

    if cell.hyperlink and cell.hyperlink.target:
        target = str(cell.hyperlink.target).strip()
        if target.lower().startswith(("http://", "https://")):
            return target

    return ""


def download_with_retry(url, timeout, retries):
    last_error = ""
    for attempt in range(retries + 1):
        try:
            data, content_type = read_bytes(url, timeout)
            if not data:
                raise ValueError("Empty response body")
            return data, content_type, ""
        except (HTTPError, URLError, TimeoutError, ValueError) as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            if attempt >= retries:
                break
    return b"", "", last_error


def main():
    args = parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: XLSX not found: {args.xlsx}")
        return 1

    image_col = column_index_from_string(args.column.upper())
    key_col = column_index_from_string(args.key_column.upper())

    output_dir = Path(args.out_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / "download_report.csv"

    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb[args.sheet] if args.sheet and args.sheet in wb.sheetnames else wb.active

    attempted = 0
    success = 0
    failed = 0
    skipped_empty_key = 0
    skipped_empty_url = 0
    skipped_exists = 0

    with report_path.open("w", newline="", encoding="utf-8") as report_file:
        writer = csv.writer(report_file)
        writer.writerow(["row", "sku", "url", "status", "file", "error"])

        for row_idx in range(args.min_row, ws.max_row + 1):
            key_value = ws.cell(row=row_idx, column=key_col).value
            if is_empty(key_value):
                skipped_empty_key += 1
                continue

            image_cell = ws.cell(row=row_idx, column=image_col)
            image_url = resolve_cell_url(image_cell)
            if not image_url:
                skipped_empty_url += 1
                writer.writerow([row_idx, key_value, "", "skipped_empty_url", "", ""])
                continue

            if args.limit and attempted >= args.limit:
                break

            attempted += 1
            sku_text = sanitize_filename(key_value)
            base_name = f"r{row_idx}_{sku_text}"

            data, content_type, error = download_with_retry(
                image_url,
                timeout=args.timeout,
                retries=args.retries,
            )

            if error:
                failed += 1
                writer.writerow([row_idx, key_value, image_url, "failed", "", error])
                continue

            extension = infer_extension(image_url, content_type)
            target_path = output_dir / f"{base_name}{extension}"

            if target_path.exists() and not args.overwrite:
                skipped_exists += 1
                writer.writerow([row_idx, key_value, image_url, "skipped_exists", str(target_path), ""])
                continue

            target_path.write_bytes(data)
            success += 1
            writer.writerow([row_idx, key_value, image_url, "downloaded", str(target_path), ""])

            if args.sleep > 0:
                time.sleep(args.sleep)

    print("=" * 72)
    print("XLSX Image Download Report")
    print("=" * 72)
    print(f"File              : {args.xlsx}")
    print(f"Sheet             : {ws.title}")
    print(f"Image column      : {args.column.upper()}")
    print(f"Rows scanned      : {ws.max_row - args.min_row + 1}")
    print(f"Attempted         : {attempted}")
    print(f"Downloaded        : {success}")
    print(f"Failed            : {failed}")
    print(f"Skipped empty key : {skipped_empty_key}")
    print(f"Skipped empty URL : {skipped_empty_url}")
    print(f"Skipped exists    : {skipped_exists}")
    print(f"Output dir        : {output_dir}")
    print(f"Report file       : {report_path}")
    print("=" * 72)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
