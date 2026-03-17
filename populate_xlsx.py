#!/usr/bin/env python3
"""
Populate SAN NENG.xlsx with product data from supported JSON sources.
Matches by SKU (Mfr Catalog No. in column B) and prioritizes Image Link in column E.
"""

import json
import os
import re
from openpyxl import load_workbook

# File paths
SOURCES_DIR = os.path.join(os.path.dirname(__file__), "sources")
JSON_CANDIDATES = [
    os.path.join(os.path.dirname(__file__), "sanneng", "coupang.json"),
    os.path.join(SOURCES_DIR, "coupang.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "phoonhuat.json"),
    os.path.join(SOURCES_DIR, "phoonhuat.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "sinarhimalaya.json"),
    os.path.join(SOURCES_DIR, "sinarhimalaya.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "redmanshop.json"),
    os.path.join(SOURCES_DIR, "redmanshop.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "mehsonline.json"),
    os.path.join(SOURCES_DIR, "mehsonline.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "sannenggroup.json"),
    os.path.join(SOURCES_DIR, "sannenggroup.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "unopan.json"),
    os.path.join(SOURCES_DIR, "unopan.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "invi.json"),
    os.path.join(SOURCES_DIR, "invi.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "tokopedia_v0.1.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "sannengvietnam.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "sannengvietnam_v0.1.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "chakawal_v0.1.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "kitchenworldthailand.json"),
    os.path.join(SOURCES_DIR, "kitchenworldthailand.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "kainan_v0.2.json"),
    os.path.join(os.path.dirname(__file__), "sanneng", "simplydifferent.json"),
    os.path.join(SOURCES_DIR, "simplydifferent.json"),
]
XLSX_FILE = os.path.join(SOURCES_DIR, "SAN NENG.xlsx")

# Column mapping (1-indexed for openpyxl)
COL_SKU = 2  # Column B - Mfr Catalog No.
COL_IMAGE_LINK = 5  # Column E - Image Link (priority)
COL_PRODUCT_NAME = 6  # Column F - Product Name
COL_PRODUCT_URL = 7  # Column G - Product URL
COL_CATEGORY = 8  # Column H - Category
COL_GALLERY_FULL = 9  # Column I - Gallery Full Image Links (comma-separated)
COL_GALLERY_THUMB = 10  # Column J - Gallery Thumbnail Links
COL_GALLERY_SRCSET = 11  # Column K - Gallery Thumb Srcsets
COL_IMAGE_COUNT = 12  # Column L - Gallery Image Count
COL_PAGE_TITLE = 13  # Column M - Product Page Title


def ensure_headers(ws):
    ws.cell(row=1, column=COL_IMAGE_LINK).value = "Image Link"
    ws.cell(row=1, column=COL_PRODUCT_NAME).value = "Product Name"
    ws.cell(row=1, column=COL_PRODUCT_URL).value = "Product URL"
    ws.cell(row=1, column=COL_CATEGORY).value = "Category"
    ws.cell(row=1, column=COL_GALLERY_FULL).value = "Gallery Full Image Links"
    ws.cell(row=1, column=COL_GALLERY_THUMB).value = "Gallery Thumbnail Links"
    ws.cell(row=1, column=COL_GALLERY_SRCSET).value = "Gallery Thumb Srcsets"
    ws.cell(row=1, column=COL_IMAGE_COUNT).value = "Gallery Image Count"
    ws.cell(row=1, column=COL_PAGE_TITLE).value = "Product Page Title"


def load_json_data(json_file):
    with open(json_file, "r", encoding="utf-8") as f:
        return json.load(f)


def resolve_json_file():
    for json_path in JSON_CANDIDATES:
        if os.path.exists(json_path):
            return json_path
    return ""


def resolve_json_files():
    return [path for path in JSON_CANDIDATES if os.path.exists(path)]


def create_sku_to_data_map(json_data):
    sku_map = {}

    # Pass 1: index explicit SKU fields first (authoritative)
    for item in json_data:
        direct_sku = normalize_sku(item.get("sku", ""))
        if direct_sku:
            sku_map.setdefault(direct_sku, item)

    # Pass 2: add inferred aliases only if key is still missing
    for item in json_data:
        candidates = extract_sku_candidates(item)
        for sku in candidates:
            sku_map.setdefault(sku, item)
    return sku_map


def normalize_sku(value):
    if not value:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(value).upper())


def find_item_by_sku(sku_map, sku):
    normalized = normalize_sku(sku)
    if normalized and normalized in sku_map:
        return sku_map[normalized]

    for part in split_sku_parts(sku):
        part_normalized = normalize_sku(part)
        if part_normalized and part_normalized in sku_map:
            return sku_map[part_normalized]


    # Fallback: search for normalized SKU as substring in normalized name/title/description fields in sku_map
    for item in sku_map.values():
        text_pool = [item.get("name", ""), item.get("title", ""), item.get("description", "")]
        for text in text_pool:
            if not text:
                continue
            norm_text = normalize_sku(text)
            if normalized and normalized in norm_text:
                return item

    # FINAL fallback: search all items in merged_json_data for normalized SKU in name/title/description
    global merged_json_data
    if 'merged_json_data' in globals():
        for item in merged_json_data:
            text_pool = [item.get("name", ""), item.get("title", ""), item.get("description", "")]
            for text in text_pool:
                if not text:
                    continue
                norm_text = normalize_sku(text)
                if normalized and normalized in norm_text:
                    return item

    return None


def split_sku_parts(value):
    if value is None:
        return []

    text = str(value).strip()
    if not text:
        return []

    # Handles combined catalog values like: SN1071/T231176, SN1071;T231176, SN1071,T231176
    parts = re.split(r"[\/|,;\n\r\t]+", text)
    return [part.strip() for part in parts if part and part.strip()]


def infer_sku_from_item(item):
    candidates = [
        item.get("name", ""),
        item.get("title", ""),
        item.get("description", ""),
        item.get("product_url", ""),
        item.get("url", ""),
    ]
    # Supports patterns like SN7086, SN-7086, SN 7086, TIP-133, TS-783
    pattern = re.compile(r"\b(?:SN|TIP|TS)[\s\-]*\d+[A-Z]?\b", re.IGNORECASE)
    for text in candidates:
        if not text:
            continue
        match = pattern.search(str(text))
        if match:
            return normalize_sku(match.group(0))

        # Also support alphanumeric catalog codes like GA-135, S-155 from descriptions
        generic_match = re.search(r"\b[A-Z]{1,5}[\s\-]*\d{2,6}[A-Z]?\b", str(text), re.IGNORECASE)
        if generic_match:
            return normalize_sku(generic_match.group(0))
    return ""


def extract_sku_candidates(item):
    sku_candidates = []

    direct_sku = normalize_sku(item.get("sku", ""))
    if direct_sku:
        sku_candidates.append(direct_sku)

    inferred = infer_sku_from_item(item)
    if inferred:
        sku_candidates.append(inferred)

    text_pool = [
        item.get("name", ""),
        item.get("title", ""),
        item.get("description", ""),
        item.get("product_url", ""),
        item.get("url", ""),
    ]

    pattern = re.compile(r"\b[A-Z]{1,5}[\s\-]*\d{2,6}[A-Z]?\b", re.IGNORECASE)
    for text in text_pool:
        if not text:
            continue
        for raw in pattern.findall(str(text)):
            normalized = normalize_sku(raw)
            if normalized:
                sku_candidates.append(normalized)

    # preserve order while removing duplicates
    unique_candidates = []
    seen = set()
    for sku in sku_candidates:
        if sku in seen:
            continue
        seen.add(sku)
        unique_candidates.append(sku)

    return unique_candidates


def format_image_link(item):
    gallery_links = extract_gallery_links(item)
    if gallery_links:
        return gallery_links[0]
    # Fallback to listing image
    return item.get("image", "") or item.get("image_url", "")


def extract_gallery_links(item):
    return (
        item.get("gallery_full_image_links")
        or item.get("gallery_images")
        or item.get("detail_image_urls")
        or []
    )


def format_gallery_list(links):
    if not links:
        return ""
    return "; ".join(links)


def populate_xlsx(xlsx_file, sku_map):
    wb = load_workbook(xlsx_file)
    ws = wb.active
    ensure_headers(ws)

    matched_count = 0
    unmatched_count = 0
    skipped_count = 0

    # Iterate through rows starting from row 2 (assuming row 1 is headers)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        sku_cell = row[COL_SKU - 1]  # Adjust for 0-indexing
        sku = sku_cell.value

        if not sku:
            skipped_count += 1
            continue

        item = find_item_by_sku(sku_map, sku)

        if item is None:
            unmatched_count += 1
            continue
        
        # Populate columns in priority order
        image_link = format_image_link(item)
        product_name = item.get("name", "")
        product_url = item.get("url", "") or item.get("product_url", "")
        category = item.get("category", "")
        gallery_full_links = extract_gallery_links(item)
        if not gallery_full_links:
            image_link_fallback = item.get("image", "") or item.get("image_url", "")
            gallery_full_links = [image_link_fallback] if image_link_fallback else []
        gallery_full = format_gallery_list(gallery_full_links)
        gallery_thumb = format_gallery_list(item.get("gallery_thumbnail_links", []))
        gallery_srcset = format_gallery_list(item.get("gallery_thumb_srcsets", []))
        image_count = item.get("gallery_image_count") or item.get("detail_image_count") or len(gallery_full_links)
        page_title = item.get("product_page_title") or item.get("page_title") or item.get("title", "")

        # Write to cells
        ws.cell(row=row_idx, column=COL_IMAGE_LINK).value = image_link
        ws.cell(row=row_idx, column=COL_PRODUCT_NAME).value = product_name
        ws.cell(row=row_idx, column=COL_PRODUCT_URL).value = product_url
        ws.cell(row=row_idx, column=COL_CATEGORY).value = category
        ws.cell(row=row_idx, column=COL_GALLERY_FULL).value = gallery_full
        ws.cell(row=row_idx, column=COL_GALLERY_THUMB).value = gallery_thumb
        ws.cell(row=row_idx, column=COL_GALLERY_SRCSET).value = gallery_srcset
        ws.cell(row=row_idx, column=COL_IMAGE_COUNT).value = image_count
        ws.cell(row=row_idx, column=COL_PAGE_TITLE).value = page_title

        matched_count += 1

    # Save the workbook
    wb.save(xlsx_file)

    return matched_count, unmatched_count, skipped_count


def main():
    print("=" * 70)
    print("SAN NENG XLSX Populator")
    print("=" * 70)

    json_files = resolve_json_files()

    # Check files exist
    if not json_files:
        print("ERROR: No supported JSON source file found.")
        print("Checked:")
        for path in JSON_CANDIDATES:
            print(f"  - {path}")
        return False

    if not os.path.exists(XLSX_FILE):
        print(f"ERROR: XLSX file not found at {XLSX_FILE}")
        return False


    # Make merged_json_data global for fallback matching
    global merged_json_data
    merged_json_data = []
    print("\nLoading JSON sources:")
    for json_file in json_files:
        source_data = load_json_data(json_file)
        merged_json_data.extend(source_data)
        print(f"  ✓ {json_file} ({len(source_data)} items)")
    print(f"  ✓ Total merged items: {len(merged_json_data)}")

    print(f"\nCreating SKU map...")
    sku_map = create_sku_to_data_map(merged_json_data)
    print(f"  ✓ Created map with {len(sku_map)} unique SKUs")

    print(f"\nPopulating XLSX: {XLSX_FILE}")
    matched, unmatched, skipped = populate_xlsx(XLSX_FILE, sku_map)

    print(f"\nResults:")
    print(f"  ✓ Matched and populated: {matched} rows")
    print(f"  ✗ Unmatched SKUs: {unmatched} rows")
    print(f"  ⊘ Skipped (empty SKU): {skipped} rows")

    print(f"\nColumn mapping:")
    print(f"  Column B: Mfr Catalog No. (SKU) - match key")
    print(f"  Column E: Image Link (PRIORITY - first populated)")
    print(f"  Column F: Product Name")
    print(f"  Column G: Product URL")
    print(f"  Column H: Category")
    print(f"  Column I: Gallery Full Image Links")
    print(f"  Column J: Gallery Thumbnail Links")
    print(f"  Column K: Gallery Thumb Srcsets")
    print(f"  Column L: Gallery Image Count")
    print(f"  Column M: Product Page Title")

    print(f"\n✓ XLSX file saved successfully!")
    print("=" * 70)

    return True


if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
