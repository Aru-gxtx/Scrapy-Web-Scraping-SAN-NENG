#!/usr/bin/env python3
"""
Report missing data in an XLSX file from a start column onward.

Default behavior:
- Reads from column E onward
- Scans data rows starting at row 2
- Includes only rows where key column B has a value
"""

import argparse
import os

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter


def is_empty(value):
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def parse_args():
    default_xlsx = os.path.join(os.path.dirname(__file__), "sources", "SAN NENG.xlsx")

    parser = argparse.ArgumentParser(description="XLSX missing-data report")
    parser.add_argument("--xlsx", default=default_xlsx, help="Path to XLSX file")
    parser.add_argument("--sheet", default="", help="Sheet name (default: active sheet)")
    parser.add_argument("--start-column", default="E", help="Start column letter (default: E)")
    parser.add_argument("--key-column", default="B", help="Only include rows where this column has value (default: B)")
    parser.add_argument("--min-row", type=int, default=2, help="First row to scan (default: 2)")
    return parser.parse_args()


def main():
    args = parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: XLSX not found: {args.xlsx}")
        return 1

    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb[args.sheet] if args.sheet and args.sheet in wb.sheetnames else wb.active

    start_col = column_index_from_string(args.start_column.upper())
    key_col = column_index_from_string(args.key_column.upper())
    max_col = ws.max_column
    max_row = ws.max_row

    if start_col > max_col:
        print(f"ERROR: Start column {args.start_column} is beyond max used column {get_column_letter(max_col)}")
        return 1

    total_rows_scanned = 0
    rows_with_missing = 0
    total_cells_expected = 0
    total_cells_missing = 0
    total_cells_filled = 0

    missing_by_column = {col_idx: 0 for col_idx in range(start_col, max_col + 1)}

    for row_idx in range(args.min_row, max_row + 1):
        key_value = ws.cell(row=row_idx, column=key_col).value
        if is_empty(key_value):
            continue

        total_rows_scanned += 1
        row_missing = 0

        for col_idx in range(start_col, max_col + 1):
            total_cells_expected += 1
            value = ws.cell(row=row_idx, column=col_idx).value
            if is_empty(value):
                total_cells_missing += 1
                missing_by_column[col_idx] += 1
                row_missing += 1
            else:
                total_cells_filled += 1

        if row_missing > 0:
            rows_with_missing += 1

    completion_pct = (total_cells_filled / total_cells_expected * 100.0) if total_cells_expected else 0.0
    missing_pct = (total_cells_missing / total_cells_expected * 100.0) if total_cells_expected else 0.0

    print("=" * 72)
    print("XLSX Missing Data Report")
    print("=" * 72)
    print(f"File              : {args.xlsx}")
    print(f"Sheet             : {ws.title}")
    print(f"Rows scanned      : {total_rows_scanned} (key column {args.key_column.upper()} not empty)")
    print(f"Columns scanned   : {args.start_column.upper()} to {get_column_letter(max_col)}")
    print(f"Cell coverage     : {total_cells_expected}")
    print(f"Filled cells      : {total_cells_filled}")
    print(f"Missing cells     : {total_cells_missing}")
    print(f"Rows with missing : {rows_with_missing}")
    print(f"Completion        : {completion_pct:.2f}%")
    print(f"Missing ratio     : {missing_pct:.2f}%")

    print("\nMissing by column:")
    for col_idx in range(start_col, max_col + 1):
        col_letter = get_column_letter(col_idx)
        miss = missing_by_column[col_idx]
        col_pct = (miss / total_rows_scanned * 100.0) if total_rows_scanned else 0.0
        print(f"  {col_letter}: {miss} rows missing ({col_pct:.2f}%)")

    print("=" * 72)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
